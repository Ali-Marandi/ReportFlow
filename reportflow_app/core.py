"""Core services for ReportFlow Desktop.

The module deliberately keeps business logic independent of the GUI so that it can
be tested, automated, and later exposed through enterprise service adapters.
"""
from __future__ import annotations

import csv
import json
import sqlite3
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable

import keyring
import matplotlib
import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

matplotlib.use("Agg")

APP_NAME = "ReportFlow"
APP_VERSION = "1.0.0"


@dataclass(slots=True)
class DataProfile:
    rows: int
    columns: int
    duplicate_rows: int
    missing_cells: int
    numeric_columns: list[str]
    categorical_columns: list[str]
    date_columns: list[str]
    warnings: list[str]


@dataclass(slots=True)
class ReportDefinition:
    id: int | None
    name: str
    source_path: str
    template: str
    title: str
    selected_columns: list[str]
    formats: list[str]
    created_at: str
    updated_at: str


@dataclass(slots=True)
class RunResult:
    run_id: int
    report_id: int | None
    status: str
    started_at: str
    finished_at: str
    artifacts: list[str]
    message: str


class ReportFlowError(Exception):
    """A user-safe error for the ReportFlow presentation layer."""


class ProjectStore:
    """SQLite-backed report catalog, scheduling metadata, and append-only audit trail."""

    def __init__(self, database_path: Path) -> None:
        self.database_path = Path(database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                PRAGMA foreign_keys = ON;
                CREATE TABLE IF NOT EXISTS reports (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    source_path TEXT NOT NULL,
                    template TEXT NOT NULL,
                    title TEXT NOT NULL,
                    selected_columns TEXT NOT NULL,
                    formats TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS report_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    report_id INTEGER,
                    status TEXT NOT NULL,
                    started_at TEXT NOT NULL,
                    finished_at TEXT NOT NULL,
                    artifacts TEXT NOT NULL,
                    message TEXT NOT NULL,
                    FOREIGN KEY(report_id) REFERENCES reports(id) ON DELETE SET NULL
                );
                CREATE TABLE IF NOT EXISTS schedules (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    report_id INTEGER NOT NULL,
                    hour INTEGER NOT NULL CHECK(hour BETWEEN 0 AND 23),
                    minute INTEGER NOT NULL CHECK(minute BETWEEN 0 AND 59),
                    enabled INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(report_id) REFERENCES reports(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS audit_events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT NOT NULL,
                    actor TEXT NOT NULL,
                    action TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    entity_id TEXT,
                    details TEXT NOT NULL
                );
                """
            )

    def audit(self, action: str, entity_type: str, entity_id: int | str | None, details: dict | None = None, actor: str = "local-user") -> None:
        with self._connect() as connection:
            connection.execute(
                "INSERT INTO audit_events(timestamp, actor, action, entity_type, entity_id, details) VALUES (?, ?, ?, ?, ?, ?)",
                (utc_now(), actor, action, entity_type, None if entity_id is None else str(entity_id), json.dumps(details or {}, ensure_ascii=False)),
            )

    def save_report(self, definition: ReportDefinition) -> ReportDefinition:
        now = utc_now()
        payload = (
            definition.name.strip(),
            str(Path(definition.source_path).expanduser()),
            definition.template,
            definition.title.strip(),
            json.dumps(definition.selected_columns, ensure_ascii=False),
            json.dumps(definition.formats, ensure_ascii=False),
        )
        if not definition.name.strip() or not definition.source_path:
            raise ReportFlowError("A report name and source file are required.")
        if not definition.formats:
            raise ReportFlowError("Choose at least one export format.")
        with self._connect() as connection:
            if definition.id is None:
                cursor = connection.execute(
                    """INSERT INTO reports(name, source_path, template, title, selected_columns, formats, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (*payload, now, now),
                )
                report_id = int(cursor.lastrowid)
                action = "report.created"
            else:
                report_id = definition.id
                connection.execute(
                    """UPDATE reports SET name=?, source_path=?, template=?, title=?, selected_columns=?, formats=?, updated_at=? WHERE id=?""",
                    (*payload, now, report_id),
                )
                action = "report.updated"
        saved = ReportDefinition(
            id=report_id,
            name=payload[0],
            source_path=payload[1],
            template=payload[2],
            title=payload[3],
            selected_columns=list(definition.selected_columns),
            formats=list(definition.formats),
            created_at=now if definition.id is None else definition.created_at,
            updated_at=now,
        )
        self.audit(action, "report", report_id, {"name": saved.name, "formats": saved.formats})
        return saved

    def list_reports(self) -> list[ReportDefinition]:
        with self._connect() as connection:
            rows = connection.execute("SELECT * FROM reports ORDER BY updated_at DESC").fetchall()
        return [self._row_to_report(row) for row in rows]

    def get_report(self, report_id: int) -> ReportDefinition:
        with self._connect() as connection:
            row = connection.execute("SELECT * FROM reports WHERE id=?", (report_id,)).fetchone()
        if row is None:
            raise ReportFlowError("The selected report no longer exists.")
        return self._row_to_report(row)

    def delete_report(self, report_id: int) -> None:
        with self._connect() as connection:
            connection.execute("DELETE FROM reports WHERE id=?", (report_id,))
        self.audit("report.deleted", "report", report_id)

    def record_run(self, result: RunResult) -> None:
        with self._connect() as connection:
            connection.execute(
                """INSERT INTO report_runs(report_id, status, started_at, finished_at, artifacts, message)
                VALUES (?, ?, ?, ?, ?, ?)""",
                (result.report_id, result.status, result.started_at, result.finished_at, json.dumps(result.artifacts), result.message),
            )
        self.audit("report.executed", "report", result.report_id, {"status": result.status, "artifacts": result.artifacts})

    def list_runs(self, limit: int = 100) -> list[sqlite3.Row]:
        with self._connect() as connection:
            return connection.execute(
                """SELECT report_runs.*, reports.name AS report_name FROM report_runs
                LEFT JOIN reports ON reports.id=report_runs.report_id
                ORDER BY report_runs.id DESC LIMIT ?""",
                (limit,),
            ).fetchall()

    def list_audit_events(self, limit: int = 200) -> list[sqlite3.Row]:
        with self._connect() as connection:
            return connection.execute("SELECT * FROM audit_events ORDER BY id DESC LIMIT ?", (limit,)).fetchall()

    def upsert_schedule(self, report_id: int, hour: int, minute: int) -> int:
        now = utc_now()
        with self._connect() as connection:
            old = connection.execute("SELECT id FROM schedules WHERE report_id=?", (report_id,)).fetchone()
            if old:
                connection.execute("UPDATE schedules SET hour=?, minute=?, enabled=1 WHERE id=?", (hour, minute, old["id"]))
                schedule_id = int(old["id"])
            else:
                cursor = connection.execute(
                    "INSERT INTO schedules(report_id, hour, minute, enabled, created_at) VALUES (?, ?, ?, 1, ?)",
                    (report_id, hour, minute, now),
                )
                schedule_id = int(cursor.lastrowid)
        self.audit("schedule.saved", "schedule", schedule_id, {"report_id": report_id, "hour": hour, "minute": minute})
        return schedule_id

    def list_schedules(self) -> list[sqlite3.Row]:
        with self._connect() as connection:
            return connection.execute(
                """SELECT schedules.*, reports.name AS report_name FROM schedules
                JOIN reports ON reports.id=schedules.report_id WHERE schedules.enabled=1 ORDER BY hour, minute"""
            ).fetchall()

    @staticmethod
    def _row_to_report(row: sqlite3.Row) -> ReportDefinition:
        return ReportDefinition(
            id=row["id"], name=row["name"], source_path=row["source_path"], template=row["template"],
            title=row["title"], selected_columns=json.loads(row["selected_columns"]), formats=json.loads(row["formats"]),
            created_at=row["created_at"], updated_at=row["updated_at"],
        )


class CredentialVault:
    """OS-backed credential access. Passwords and tokens are never stored in project files."""

    SERVICE = "ReportFlow"

    @classmethod
    def set_secret(cls, reference: str, value: str) -> None:
        if not reference.strip() or not value:
            raise ReportFlowError("A credential reference and secret value are required.")
        keyring.set_password(cls.SERVICE, reference.strip(), value)

    @classmethod
    def get_secret(cls, reference: str) -> str | None:
        return keyring.get_password(cls.SERVICE, reference.strip())

    @classmethod
    def delete_secret(cls, reference: str) -> None:
        try:
            keyring.delete_password(cls.SERVICE, reference.strip())
        except keyring.errors.PasswordDeleteError:
            pass


class DataInspector:
    ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

    @classmethod
    def load(cls, path: str | Path) -> pd.DataFrame:
        file_path = Path(path).expanduser()
        if not file_path.exists():
            raise ReportFlowError("The selected source file cannot be found.")
        if file_path.suffix.lower() not in cls.ALLOWED_EXTENSIONS:
            raise ReportFlowError("Only CSV and Excel files are supported in this release.")
        try:
            if file_path.suffix.lower() == ".csv":
                return pd.read_csv(file_path)
            return pd.read_excel(file_path)
        except Exception as error:  # pandas provides varied engine-specific errors
            raise ReportFlowError(f"Unable to read the source file: {error}") from error

    @staticmethod
    def profile(data: pd.DataFrame) -> DataProfile:
        if data.empty:
            return DataProfile(0, len(data.columns), 0, 0, [], list(data.columns), [], ["The source file contains no data rows."])
        numeric = data.select_dtypes(include="number").columns.tolist()
        date_columns: list[str] = []
        for column in data.columns:
            if pd.api.types.is_datetime64_any_dtype(data[column]):
                date_columns.append(str(column))
        categorical = [str(column) for column in data.columns if str(column) not in numeric and str(column) not in date_columns]
        warnings: list[str] = []
        duplicates = int(data.duplicated().sum())
        missing = int(data.isna().sum().sum())
        if duplicates:
            warnings.append(f"{duplicates:,} duplicate rows detected.")
        if missing:
            warnings.append(f"{missing:,} blank cells detected.")
        if not numeric:
            warnings.append("No numeric columns were detected; visual summaries will be limited.")
        return DataProfile(len(data), len(data.columns), duplicates, missing, numeric, categorical, date_columns, warnings)

    @staticmethod
    def sample_data() -> pd.DataFrame:
        return pd.DataFrame(
            {
                "Period": ["Q1", "Q2", "Q3", "Q4"],
                "Revenue": [125000, 142500, 158000, 184000],
                "Operating Cost": [80500, 88400, 95750, 101200],
                "Net Profit": [28400, 35100, 42250, 53900],
                "Customer Satisfaction": [87, 89, 91, 93],
            }
        )


class ReportRenderer:
    """Creates branded HTML, PDF, and Excel deliverables from a tabular data source."""

    BRAND_NAVY = "#102A43"
    BRAND_BLUE = "#0F6CBD"
    BRAND_TEAL = "#12B886"
    BRAND_SLATE = "#486581"

    def __init__(self, export_directory: Path) -> None:
        self.export_directory = Path(export_directory)
        self.export_directory.mkdir(parents=True, exist_ok=True)

    def render(self, report: ReportDefinition, data: pd.DataFrame) -> list[str]:
        if data.empty:
            raise ReportFlowError("A report cannot be generated from an empty dataset.")
        selected = [column for column in report.selected_columns if column in data.columns]
        frame = data[selected] if selected else data.copy()
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        stem = safe_file_stem(report.name) + "-" + timestamp
        artifacts: list[str] = []
        chart_path = self._create_chart(frame, stem)
        for fmt in report.formats:
            normalized = fmt.lower()
            if normalized == "html":
                artifacts.append(str(self._render_html(report, frame, chart_path, stem)))
            elif normalized == "pdf":
                artifacts.append(str(self._render_pdf(report, frame, chart_path, stem)))
            elif normalized in {"xlsx", "excel"}:
                artifacts.append(str(self._render_excel(report, frame, stem)))
        return artifacts

    def _create_chart(self, frame: pd.DataFrame, stem: str) -> Path:
        numeric = frame.select_dtypes(include="number").columns.tolist()
        chart_path = self.export_directory / f"{stem}-insight.png"
        figure = Figure(figsize=(9, 4.6), facecolor="white")
        axis = figure.add_subplot(111)
        if numeric:
            x_values = range(len(frame))
            for column in numeric[:4]:
                axis.plot(x_values, frame[column].fillna(0), marker="o", linewidth=2.4, label=str(column))
            axis.set_xticks(list(x_values))
            first_column = frame.columns[0]
            axis.set_xticklabels(frame[first_column].astype(str).tolist(), rotation=0)
            axis.legend(frameon=False, ncol=min(len(numeric), 4), loc="upper left")
            axis.set_title("Performance trend", loc="left", fontsize=15, fontweight="bold", color=self.BRAND_NAVY)
            axis.grid(axis="y", alpha=0.18)
        else:
            counts = frame.iloc[:, 0].astype(str).value_counts().head(10)
            axis.bar(counts.index, counts.values, color=self.BRAND_BLUE)
            axis.set_title("Source distribution", loc="left", fontsize=15, fontweight="bold", color=self.BRAND_NAVY)
            axis.tick_params(axis="x", rotation=25)
        figure.tight_layout()
        figure.savefig(chart_path, dpi=160, bbox_inches="tight")
        return chart_path

    def _render_html(self, report: ReportDefinition, frame: pd.DataFrame, chart_path: Path, stem: str) -> Path:
        profile = DataInspector.profile(frame)
        output = self.export_directory / f"{stem}.html"
        kpis = self._kpi_cards(frame)
        rows = frame.head(250).to_html(index=False, classes="data-table", border=0, escape=True)
        html = f"""<!doctype html>
<html lang=\"en\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">
<title>{escape_html(report.title)}</title><style>
:root{{--navy:{self.BRAND_NAVY};--blue:{self.BRAND_BLUE};--muted:{self.BRAND_SLATE};--line:#D9E2EC;}}
*{{box-sizing:border-box}} body{{margin:0;background:#F5F7FA;color:var(--navy);font-family:Inter,Segoe UI,Arial,sans-serif}} .page{{max-width:1180px;margin:32px auto;background:#fff;padding:42px;border-radius:20px;box-shadow:0 12px 40px #102a4314}} .eyebrow{{color:var(--blue);letter-spacing:1.4px;font-size:12px;font-weight:700;text-transform:uppercase}} h1{{font-size:32px;margin:8px 0}} .meta{{color:var(--muted);margin:0 0 28px}} .kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:22px 0 30px}} .kpi{{border:1px solid var(--line);border-radius:14px;padding:18px;background:#fff}} .kpi-label{{color:var(--muted);font-size:12px}} .kpi-value{{font-size:24px;font-weight:750;margin-top:6px}} .section{{margin-top:32px}} h2{{font-size:18px;margin:0 0 14px}} .chart{{width:100%;border:1px solid var(--line);border-radius:12px;padding:8px}} .data-table{{border-collapse:collapse;width:100%;font-size:13px}} .data-table th{{background:var(--navy);color:white;text-align:left;padding:10px}} .data-table td{{border-bottom:1px solid var(--line);padding:9px}} .foot{{color:var(--muted);font-size:12px;margin-top:32px;border-top:1px solid var(--line);padding-top:16px}} @media(max-width:700px){{.page{{margin:0;border-radius:0;padding:24px}}.kpis{{grid-template-columns:repeat(2,1fr)}}}}</style></head>
<body><main class=\"page\"><div class=\"eyebrow\">ReportFlow · {escape_html(report.template)} report</div><h1>{escape_html(report.title)}</h1><p class=\"meta\">Generated {datetime.now().strftime('%d %b %Y, %H:%M')} · {profile.rows:,} records · {profile.columns} fields</p><section class=\"kpis\">{kpis}</section><section class=\"section\"><h2>Executive insight</h2><img class=\"chart\" src=\"{chart_path.name}\" alt=\"Report insight chart\"></section><section class=\"section\"><h2>Validated source detail</h2>{rows}</section><div class=\"foot\">Produced by ReportFlow Desktop · Audit-ready local generation</div></main></body></html>"""
        output.write_text(html, encoding="utf-8")
        return output

    def _render_pdf(self, report: ReportDefinition, frame: pd.DataFrame, chart_path: Path, stem: str) -> Path:
        output = self.export_directory / f"{stem}.pdf"
        document = SimpleDocTemplate(str(output), pagesize=landscape(A4), rightMargin=1.1 * cm, leftMargin=1.1 * cm, topMargin=1.1 * cm, bottomMargin=1.1 * cm)
        styles = getSampleStyleSheet()
        styles["Title"].textColor = colors.HexColor(self.BRAND_NAVY)
        styles["Heading2"].textColor = colors.HexColor(self.BRAND_NAVY)
        story = [Paragraph(report.title, styles["Title"]), Spacer(1, 0.18 * cm)]
        story.append(Paragraph(f"{report.template} report · Generated {datetime.now().strftime('%d %b %Y, %H:%M')}", styles["Normal"]))
        story.extend([Spacer(1, 0.35 * cm), Image(str(chart_path), width=23.5 * cm, height=10.5 * cm), Spacer(1, 0.35 * cm), Paragraph("Validated source detail", styles["Heading2"]), Spacer(1, 0.16 * cm)])
        preview = frame.head(28).copy()
        values = [list(map(str, preview.columns))] + [[format_cell(item) for item in row] for row in preview.itertuples(index=False, name=None)]
        table = Table(values, repeatRows=1, colWidths=[25.7 * cm / max(len(preview.columns), 1)] * len(preview.columns))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(self.BRAND_NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.2),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E2EC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        document.build(story)
        return output

    def _render_excel(self, report: ReportDefinition, frame: pd.DataFrame, stem: str) -> Path:
        output = self.export_directory / f"{stem}.xlsx"
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Overview"
        overview["A1"] = report.title
        overview["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        overview["A1"].fill = PatternFill("solid", fgColor="102A43")
        overview.merge_cells("A1:F1")
        overview["A3"] = "Template"
        overview["B3"] = report.template
        overview["A4"] = "Generated"
        overview["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        overview["A5"] = "Rows"
        overview["B5"] = len(frame)
        overview["A6"] = "Columns"
        overview["B6"] = len(frame.columns)
        for key in ["A3", "A4", "A5", "A6"]:
            overview[key].font = Font(bold=True, color="0F6CBD")
        sheet = workbook.create_sheet("Data")
        for column_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(1, column_index, str(column))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="102A43")
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row_index, column_index, value if not pd.isna(value) else "")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 35)
            sheet.column_dimensions[column[0].column_letter].width = width
        numeric_columns = [index for index, dtype in enumerate(frame.dtypes, start=1) if pd.api.types.is_numeric_dtype(dtype)]
        if numeric_columns and len(frame) > 1:
            chart = BarChart()
            chart.title = "Performance overview"
            chart.y_axis.title = "Value"
            chart.x_axis.title = str(frame.columns[0])
            values = Reference(sheet, min_col=numeric_columns[0], min_row=1, max_row=min(len(frame) + 1, 40))
            categories = Reference(sheet, min_col=1, min_row=2, max_row=min(len(frame) + 1, 40))
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 16
            overview.add_chart(chart, "A9")
        workbook.save(output)
        return output

    def _kpi_cards(self, frame: pd.DataFrame) -> str:
        profile = DataInspector.profile(frame)
        kpis: list[tuple[str, str]] = [("Records", f"{profile.rows:,}"), ("Fields", str(profile.columns)), ("Data completeness", f"{(1 - profile.missing_cells / max(profile.rows * max(profile.columns, 1), 1)) * 100:.1f}%")]
        numeric = frame.select_dtypes(include="number")
        if not numeric.empty:
            kpis.append((f"Total {numeric.columns[0]}", f"{numeric.iloc[:, 0].sum():,.2f}"))
        else:
            kpis.append(("Quality alerts", str(len(profile.warnings))))
        return "".join(f'<div class="kpi"><div class="kpi-label">{escape_html(label)}</div><div class="kpi-value">{escape_html(value)}</div></div>' for label, value in kpis)


class ReportService:
    def __init__(self, store: ProjectStore, renderer: ReportRenderer) -> None:
        self.store = store
        self.renderer = renderer

    def execute(self, definition: ReportDefinition) -> RunResult:
        started = utc_now()
        run_id = int(datetime.now().timestamp() * 1000)
        try:
            data = DataInspector.load(definition.source_path)
            artifacts = self.renderer.render(definition, data)
            result = RunResult(run_id, definition.id, "completed", started, utc_now(), artifacts, "Report generated successfully.")
        except Exception as error:
            result = RunResult(run_id, definition.id, "failed", started, utc_now(), [], str(error))
        self.store.record_run(result)
        if result.status == "failed":
            raise ReportFlowError(result.message)
        return result


class ReportScheduler:
    """Local scheduled execution. Schedules are reloaded from the project database on startup."""

    def __init__(self, store: ProjectStore, service: ReportService) -> None:
        self.store = store
        self.service = service
        # Let APScheduler select the host's local timezone. Converting the Windows
        # timezone display name to a string can produce a non-IANA value such as
        # "Coordinated Universal Time", which is not portable across runners.
        self.scheduler = BackgroundScheduler()

    def start(self) -> None:
        if not self.scheduler.running:
            self.scheduler.start()
        self.reload()

    def reload(self) -> None:
        self.scheduler.remove_all_jobs()
        for item in self.store.list_schedules():
            self.scheduler.add_job(self._run_scheduled, "cron", id=f"report-{item['report_id']}", replace_existing=True, hour=item["hour"], minute=item["minute"], args=[int(item["report_id"])])

    def _run_scheduled(self, report_id: int) -> None:
        try:
            self.service.execute(self.store.get_report(report_id))
        except Exception as error:
            self.store.audit("schedule.failed", "report", report_id, {"error": str(error)})

    def shutdown(self) -> None:
        if self.scheduler.running:
            self.scheduler.shutdown(wait=False)


def utc_now() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def safe_file_stem(value: str) -> str:
    allowed = "-_.() "
    cleaned = "".join(character for character in value.strip() if character.isalnum() or character in allowed).strip().replace(" ", "-")
    return cleaned[:80] or "report"


def escape_html(value: object) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def format_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)
